"""Excel inventory import service — parses TOOLOR Excel format → product dicts."""

import io
import re
import uuid
from collections import defaultdict
from decimal import Decimal

import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

# Sheet name (from Excel) → our subcategory
SHEET_TO_SUBCAT = {
    "2667 Рубашки для взрослых": "Рубашки",
    "2615 Брюки для взрослых": "Брюки",
    "2617 Брюки спортивные для взрос": "Спортивные брюки",
    "2637 Кардиганы для взрослых": "Кардиганы",
    "2638 Лонгсливы для взрослых": "Лонгсливы",
    "2643 Свитшоты для взрослых": "Свитшоты",
    "2646 Толстовки для взрослых": "Толстовки",
    "2628 Костюмы спортивные для взр": "Спортивные костюмы",
    "2573 Ветровки для взрослых": "Ветровки",
    "2576 Куртки для взрослых": "Куртки",
    "2585 Плащи для взрослых": "Плащи",
    "2591 Жилеты утепленные для взро": "Жилеты",
    "2663 Футболки для взрослых": "Футболки",
    "2270 Пиджаки": "Пиджаки",
}

_RU_TRANSLIT = {
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'zh','з':'z','и':'i',
    'й':'j','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t',
    'у':'u','ф':'f','х':'h','ц':'ts','ч':'ch','ш':'sh','щ':'sch','ъ':'','ы':'y','ь':'',
    'э':'e','ю':'yu','я':'ya',
}


def slugify(s: str) -> str:
    s = s.lower().strip()
    s = ''.join(_RU_TRANSLIT.get(c, c) for c in s)
    s = re.sub(r'[^\w\s-]', '', s)
    s = re.sub(r'[\s_]+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s or "product"


def strip_variant_suffix(name: str) -> str:
    return re.sub(r'\s*\([^)]*\)\s*$', '', name).strip()


def extract_color_size_from_name(name: str) -> tuple[str, str]:
    match = re.search(r'\(([^)]*)\)\s*$', name)
    if not match:
        return "", ""
    parts = [p.strip() for p in match.group(1).split(",")]
    size = ""
    color_parts = []
    for p in parts:
        if re.match(r'^(XS|S|M|L|XL|XXL|XXXL|2XL|3XL|4XL|5XL|\d{2})$', p, re.IGNORECASE):
            size = p.upper()
        else:
            color_parts.append(p)
    color = " ".join(color_parts).strip().capitalize()
    return color, size


def map_gender_to_category(gender: str | None, name: str = "") -> str:
    g = (gender or "").lower()
    n = (name or "").lower()
    if "жен" in g or "female" in g or "women" in g:
        return "Женщинам"
    if "муж" in g or "male" in g or "men" in g:
        return "Мужчинам"
    if re.search(r'\bжен\.?\b|женск|для девушек', n):
        return "Женщинам"
    if re.search(r'\bмуж\.?\b|мужск|для мужчин', n):
        return "Мужчинам"
    return "Аксессуары"


def parse_excel_bytes(data: bytes) -> list[dict]:
    """Parse Excel file bytes → merged product dicts."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    merge_groups: dict[str, list[dict]] = defaultdict(list)

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_SUBCAT:
            continue

        subcat = SHEET_TO_SUBCAT[sheet_name]
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]

        def col_exact(*candidates: str) -> int | None:
            for candidate in candidates:
                for i, h in enumerate(headers):
                    if h and str(h).strip().lower() == candidate.lower():
                        return i
            return None

        def col_contains(fragment: str) -> int | None:
            for i, h in enumerate(headers):
                if h and fragment.lower() in str(h).lower():
                    return i
            return None

        sku_col = col_exact("Уникальный идентификатор товара")
        name_col = col_exact("Наименование товара")
        desc_col = col_exact("Описание товара")
        price_col = col_exact("Цена товара")
        discount_col = col_exact("Сумма скидки")
        images_col = col_contains("ссылки на изображения")
        merge_col = col_contains("объединить карточки")
        size_col = col_exact("Размер")
        color_col = col_exact("Название цвета от производителя")
        gender_col = col_exact("Пол")
        stock_col = col_contains("азия молл")

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_name = row[name_col]
            if not raw_name or str(raw_name).strip().lower() in ("nan", "none", "", "null"):
                continue
            cleaned_name = re.sub(r'\s+', ' ', str(raw_name)).strip()
            if not cleaned_name or cleaned_name.lower() in ("nan", "none", "null"):
                continue
            if cleaned_name.lower().startswith("nan "):
                cleaned_name = subcat + cleaned_name[3:]

            sku = str(row[sku_col]).strip() if sku_col is not None and row[sku_col] else None
            name = cleaned_name
            merge_raw = str(row[merge_col]).strip() if merge_col is not None and row[merge_col] else ""

            if merge_raw:
                skus_in_group = sorted([s.strip() for s in merge_raw.split(",") if s.strip()])
                canonical_key = skus_in_group[0] if skus_in_group else name
            else:
                # Fallback: group by base name + color (from name suffix) within same sheet
                base = strip_variant_suffix(name)
                name_color, _ = extract_color_size_from_name(name)
                canonical_key = f"{sheet_name}::{base}::{name_color}"

            try:
                price = Decimal(str(row[price_col])) if price_col is not None and row[price_col] else Decimal("0")
            except Exception:
                price = Decimal("0")

            try:
                discount = Decimal(str(row[discount_col])) if discount_col is not None and row[discount_col] else Decimal("0")
            except Exception:
                discount = Decimal("0")

            images_str = str(row[images_col]) if images_col is not None and row[images_col] else ""
            images = [img.strip() for img in images_str.split(",") if img.strip().startswith("http")]

            try:
                stock = int(row[stock_col]) if stock_col is not None and row[stock_col] else 0
            except Exception:
                stock = 0

            size_val = str(row[size_col]).strip() if size_col is not None and row[size_col] else ""
            color_val = str(row[color_col]).strip() if color_col is not None and row[color_col] else ""
            if not size_val or not color_val:
                name_color, name_size = extract_color_size_from_name(name)
                if not size_val:
                    size_val = name_size
                if not color_val:
                    color_val = name_color

            merge_groups[canonical_key].append({
                "sku": sku,
                "name": name,
                "description": str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else "",
                "price": price,
                "discount": discount,
                "images": images,
                "size": size_val,
                "color": color_val,
                "gender": str(row[gender_col]).strip() if gender_col is not None and row[gender_col] else "",
                "stock": stock,
                "subcategory": subcat,
            })

    products = []
    for composite_key, variants in merge_groups.items():
        first = variants[0]
        base_stripped = strip_variant_suffix(first["name"])
        primary_color = next((v["color"] for v in variants if v["color"]), "")
        base_name = f"{base_stripped} ({primary_color})" if primary_color else base_stripped

        sizes = sorted(set(v["size"] for v in variants if v["size"]))
        colors = sorted(set(v["color"] for v in variants if v["color"]))
        all_images = []
        for v in variants:
            for img in v["images"]:
                if img not in all_images:
                    all_images.append(img)

        total_stock = sum(v["stock"] for v in variants)
        max_price = max(v["price"] for v in variants)
        max_discount = max(v["discount"] for v in variants)
        original_price = None
        final_price = max_price
        if max_discount > 0 and max_discount < max_price:
            original_price = max_price
            final_price = max_price - max_discount

        descriptions = [v["description"] for v in variants if v["description"]]
        description = re.sub(r'\s+', ' ', max(descriptions, key=len)).strip()[:1500] if descriptions else ""

        gender = first["gender"] or next((v["gender"] for v in variants if v["gender"]), "")
        category = map_gender_to_category(gender, base_name)

        products.append({
            "sku": first["sku"],
            "name": base_name,
            "slug": slugify(f"{base_name}-{composite_key}")[:80],
            "description": description,
            "price": final_price,
            "original_price": original_price,
            "images": all_images,
            "image_url": all_images[0] if all_images else "",
            "sizes": sizes,
            "colors": colors,
            "stock": total_stock,
            "category": category,
            "subcategory": first["subcategory"],
        })

    return products


async def import_to_db(db: AsyncSession, products: list[dict], replace_all: bool = True) -> dict:
    """Write products to database. Returns summary dict."""
    if replace_all:
        await db.execute(text("DELETE FROM store_inventory"))
        await db.execute(text("DELETE FROM order_items"))
        await db.execute(text("DELETE FROM cart_items"))
        await db.execute(text("DELETE FROM favorites"))
        await db.execute(text("DELETE FROM products"))
        await db.execute(text("DELETE FROM subcategories"))
        await db.execute(text("DELETE FROM categories"))

    # Seed inventory across ALL active stores
    stores_result = await db.execute(
        text("SELECT id FROM locations WHERE is_active = true AND type = 'store' ORDER BY sort_order")
    )
    store_ids = [row[0] for row in stores_result.fetchall()]

    cat_ids: dict[str, uuid.UUID] = {}
    subcat_ids: dict[str, uuid.UUID] = {}
    for p in products:
        if p["category"] not in cat_ids:
            cat_ids[p["category"]] = uuid.uuid4()
        key = f"{p['category']}::{p['subcategory']}"
        if key not in subcat_ids:
            subcat_ids[key] = uuid.uuid4()

    cat_order = {"Мужчинам": 1, "Женщинам": 2, "Аксессуары": 3}
    for name, cid in cat_ids.items():
        await db.execute(
            text("INSERT INTO categories (id, name, slug, sort_order) VALUES (:id, :name, :slug, :ord)"),
            {"id": cid, "name": name, "slug": slugify(name), "ord": cat_order.get(name, 10)},
        )

    for i, (key, sid) in enumerate(subcat_ids.items()):
        cat_name, subcat_name = key.split("::", 1)
        await db.execute(
            text("INSERT INTO subcategories (id, category_id, name, slug, sort_order) VALUES (:id, :cid, :name, :slug, :ord)"),
            {"id": sid, "cid": cat_ids[cat_name], "name": subcat_name,
             "slug": slugify(f"{cat_name}-{subcat_name}"), "ord": i + 1},
        )

    seen_slugs: set[str] = set()
    seen_skus: set[str] = set()
    inserted = 0
    for p in products:
        if p["price"] <= 0:
            continue

        slug = p["slug"]
        base = slug
        counter = 1
        while slug in seen_slugs:
            counter += 1
            slug = f"{base}-{counter}"
        seen_slugs.add(slug)

        sku = p["sku"]
        if sku and sku in seen_skus:
            sku = None
        elif sku:
            seen_skus.add(sku)

        cat_id = cat_ids[p["category"]]
        subcat_id = subcat_ids[f"{p['category']}::{p['subcategory']}"]
        prod_id = uuid.uuid4()

        import json
        await db.execute(
            text("""INSERT INTO products
                (id, sku, name, slug, description, price, original_price,
                 category_id, subcategory_id, image_url, images, sizes, colors,
                 stock, is_active, is_featured, sort_order)
                VALUES (:id, :sku, :name, :slug, :desc, :price, :orig,
                        :cat, :subcat, :img, CAST(:imgs AS jsonb), CAST(:sizes AS jsonb), CAST(:colors AS jsonb),
                        :stock, true, :featured, 0)"""),
            {"id": prod_id, "sku": sku, "name": p["name"], "slug": slug, "desc": p["description"],
             "price": p["price"], "orig": p["original_price"],
             "cat": cat_id, "subcat": subcat_id,
             "img": p["image_url"], "imgs": json.dumps(p["images"]),
             "sizes": json.dumps(p["sizes"]), "colors": json.dumps(p["colors"]),
             "stock": p["stock"], "featured": p["original_price"] is not None},
        )
        inserted += 1

        if store_ids and p["sizes"] and p["stock"] > 0:
            per_size = max(1, p["stock"] // len(p["sizes"]))
            for loc_id in store_ids:
                for size in p["sizes"]:
                    await db.execute(
                        text("""INSERT INTO store_inventory (id, location_id, product_id, size, quantity)
                                VALUES (:id, :loc, :prod, :size, :qty)
                                ON CONFLICT (location_id, product_id, size) DO UPDATE SET quantity = EXCLUDED.quantity"""),
                        {"id": uuid.uuid4(), "loc": loc_id, "prod": prod_id, "size": size, "qty": per_size},
                    )

    await db.commit()

    return {
        "products_inserted": inserted,
        "categories_created": len(cat_ids),
        "subcategories_created": len(subcat_ids),
    }
