"""
Import products from TOOLOR Excel inventory file.

Groups variant rows into products via "Объединить карточки в одну по SKU" column.
Maps sheet names to subcategories, gender to category.
"""

import asyncio
import json
import re
import uuid
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import asyncpg
import openpyxl

DATABASE_URL = "postgresql://neondb_owner:npg_3NKgvsFrW5io@ep-blue-boat-amiwae15-pooler.c-5.us-east-1.aws.neon.tech/neondb?sslmode=require"
EXCEL_PATH = Path(__file__).parent.parent / "INVENTORY" / "Toolor_2026.03.23_1.xlsx"

# ── Sheet → Subcategory mapping ──────────────────────────────────────────
SHEET_TO_SUBCAT = {
    "2667 Рубашки для взрослых": "Рубашки",
    "2615 Брюки для взрослых": "Брюки",
    "2617 Брюки спортивные для взрос": "Спортивные брюки",
    "2637 Кардиганы для взрослых": "Кардиганы",
    "2638 Лонгсливы для взрослых": "Лонгсливы",
    "2643 Свитшоты для взрослых": "Свитшоты",
    "2646 Толстовки для взрослых": "Толстовки",
    "2628 Костюмы спортивные для взр": "Спортивные костюмы",
    "2573 Ветровки для взрослых": "Ветровки",
    "2576 Куртки для взрослых": "Куртки",
    "2585 Плащи для взрослых": "Плащи",
    "2591 Жилеты утепленные для взро": "Жилеты",
    "2663 Футболки для взрослых": "Футболки",
    "2270 Пиджаки": "Пиджаки",
}


def slugify(text: str) -> str:
    # Transliterate Russian to Latin
    trans = {
        'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'zh','з':'z','и':'i',
        'й':'j','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t',
        'у':'u','ф':'f','х':'h','ц':'ts','ч':'ch','ш':'sh','щ':'sch','ъ':'','ы':'y','ь':'',
        'э':'e','ю':'yu','я':'ya',
    }
    text = text.lower().strip()
    text = ''.join(trans.get(c, c) for c in text)
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text).strip('-')
    return text or "product"


def strip_variant_suffix(name: str) -> str:
    """Remove '(Цвет, Размер)' suffix from product name."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', name).strip()


def extract_color_size_from_name(name: str) -> tuple[str, str]:
    """Extract color and size from '(Color, Size)' suffix."""
    match = re.search(r'\(([^)]*)\)\s*$', name)
    if not match:
        return "", ""
    parts = [p.strip() for p in match.group(1).split(",")]
    # Usually last part is size (XS/S/M/L/XL/XXL/3XL or numeric)
    size = ""
    color_parts = []
    for p in parts:
        if re.match(r'^(XS|S|M|L|XL|XXL|XXXL|2XL|3XL|4XL|5XL|\d{2})$', p, re.IGNORECASE):
            size = p.upper()
        else:
            color_parts.append(p)
    color = " ".join(color_parts).strip().capitalize()
    return color, size


def map_gender_to_category(gender: str | None, name: str = "") -> str:
    g = (gender or "").lower()
    n = (name or "").lower()

    if "жен" in g or "female" in g or "women" in g:
        return "Женщинам"
    if "муж" in g or "male" in g or "men" in g:
        return "Мужчинам"

    # Fallback: infer from product name
    if re.search(r'\bжен\.?\b|женск|\bwomen\b|для девушек', n):
        return "Женщинам"
    if re.search(r'\bмуж\.?\b|мужск|\bmen\b|для мужчин', n):
        return "Мужчинам"

    return "Аксессуары"


def parse_excel() -> list[dict]:
    """Parse Excel → list of merged products."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Raw rows grouped by merge-key (set of all SKUs that should be merged)
    # We use frozenset of SKUs as the grouping key so variants converge
    merge_groups: dict[str, list[dict]] = defaultdict(list)

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_SUBCAT:
            continue

        subcat = SHEET_TO_SUBCAT[sheet_name]
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]

        # Find column indexes by header name (exact match preferred)
        def col_exact(*candidates: str) -> int | None:
            for candidate in candidates:
                for i, h in enumerate(headers):
                    if h and str(h).strip().lower() == candidate.lower():
                        return i
            return None

        def col_contains(header_contains: str) -> int | None:
            for i, h in enumerate(headers):
                if h and header_contains.lower() in str(h).lower():
                    return i
            return None

        sku_col = col_exact("Уникальный идентификатор товара")
        name_col = col_exact("Наименование товара")
        desc_col = col_exact("Описание товара")
        price_col = col_exact("Цена товара")
        discount_col = col_exact("Сумма скидки")
        images_col = col_contains("ссылки на изображения")
        merge_col = col_contains("объединить карточки")
        size_col = col_exact("Размер")
        color_col = col_exact("Название цвета от производителя")
        gender_col = col_exact("Пол")
        stock_col = col_contains("азия молл")

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_name = row[name_col]
            if not raw_name or str(raw_name).strip().lower() in ("nan", "none", "", "null"):
                continue
            # Clean whitespace early
            import re as _re
            cleaned_name = _re.sub(r'\s+', ' ', str(raw_name)).strip()
            if not cleaned_name or cleaned_name.lower() in ("nan", "none", "null"):
                continue
            # Fix Excel data entry: names like "nan (Color, Size)" — replace nan with subcategory name
            if cleaned_name.lower().startswith("nan "):
                cleaned_name = subcat + cleaned_name[3:]

            sku = str(row[sku_col]).strip() if sku_col is not None and row[sku_col] else None
            name = cleaned_name
            merge_raw = str(row[merge_col]).strip() if merge_col is not None and row[merge_col] else ""

            if merge_raw:
                skus_in_group = sorted([s.strip() for s in merge_raw.split(",") if s.strip()])
                canonical_key = skus_in_group[0] if skus_in_group else name
            else:
                base = strip_variant_suffix(name)
                name_color, _ = extract_color_size_from_name(name)
                canonical_key = f"{sheet_name}::{base}::{name_color}"

            try:
                price = Decimal(str(row[price_col])) if price_col is not None and row[price_col] else Decimal("0")
            except Exception:
                price = Decimal("0")

            try:
                discount = Decimal(str(row[discount_col])) if discount_col is not None and row[discount_col] else Decimal("0")
            except Exception:
                discount = Decimal("0")

            images_str = str(row[images_col]) if images_col is not None and row[images_col] else ""
            images = [img.strip() for img in images_str.split(",") if img.strip().startswith("http")]

            try:
                stock = int(row[stock_col]) if stock_col is not None and row[stock_col] else 0
            except Exception:
                stock = 0

            size_val = str(row[size_col]).strip() if size_col is not None and row[size_col] else ""
            color_val = str(row[color_col]).strip() if color_col is not None and row[color_col] else ""

            # Fallback: extract from name "(Color, Size)" if columns are empty
            if not size_val or not color_val:
                name_color, name_size = extract_color_size_from_name(name)
                if not size_val:
                    size_val = name_size
                if not color_val:
                    color_val = name_color

            merge_groups[canonical_key].append({
                "sku": sku,
                "name": name,
                "description": str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else "",
                "price": price,
                "discount": discount,
                "images": images,
                "size": size_val,
                "color": color_val,
                "gender": str(row[gender_col]).strip() if gender_col is not None and row[gender_col] else "",
                "stock": stock,
                "subcategory": subcat,
            })

    # Second-level merge: collapse color variants with same base name + subcategory into ONE product
    # E.g. "Жен. брюки трубы (Бордовый)" + "(Кофе)" + "(Оливковый)" → 1 product with 3 colors
    product_groups: dict[str, list[dict]] = defaultdict(list)
    for composite_key, variants in merge_groups.items():
        first = variants[0]
        base_stripped = strip_variant_suffix(first["name"]).strip()
        # Normalize for grouping: lowercase, collapse whitespace
        norm_key = f"{first['subcategory']}::{re.sub(r'\\s+', ' ', base_stripped.lower())}"
        product_groups[norm_key].extend(variants)

    products = []
    for norm_key, all_variants in product_groups.items():
        first = all_variants[0]
        base_name = strip_variant_suffix(first["name"]).strip()

        sizes = sorted(set(v["size"] for v in all_variants if v["size"]))
        colors = sorted(set(v["color"] for v in all_variants if v["color"]))
        all_images = []
        for v in all_variants:
            for img in v["images"]:
                if img not in all_images:
                    all_images.append(img)

        total_stock = sum(v["stock"] for v in all_variants)
        max_price = max(v["price"] for v in all_variants)
        max_discount = max(v["discount"] for v in all_variants)

        original_price = None
        final_price = max_price
        if max_discount > 0 and max_discount < max_price:
            original_price = max_price
            final_price = max_price - max_discount

        descriptions = [v["description"] for v in all_variants if v["description"]]
        description = max(descriptions, key=len) if descriptions else ""
        description = re.sub(r'\s+', ' ', description).strip()[:1500]

        gender = first["gender"] or next((v["gender"] for v in all_variants if v["gender"]), "")
        category = map_gender_to_category(gender, base_name)

        products.append({
            "sku": first["sku"],
            "name": base_name,
            "slug": slugify(f"{base_name}-{norm_key}")[:80],
            "description": description,
            "price": final_price,
            "original_price": original_price,
            "images": all_images,
            "image_url": all_images[0] if all_images else "",
            "sizes": sizes,
            "colors": colors,
            "stock": total_stock,
            "category": category,
            "subcategory": first["subcategory"],
        })

    return products


async def seed(products: list[dict]):
    conn = await asyncpg.connect(DATABASE_URL, ssl="require")

    try:
        print("Clearing old data...")
        await conn.execute("DELETE FROM store_inventory")
        await conn.execute("DELETE FROM order_items")
        await conn.execute("DELETE FROM cart_items")
        await conn.execute("DELETE FROM favorites")
        await conn.execute("DELETE FROM products")
        await conn.execute("DELETE FROM subcategories")
        await conn.execute("DELETE FROM categories")

        # Find all store locations to seed inventory across
        store_rows = await conn.fetch("SELECT id, name FROM locations WHERE is_active = true AND type = 'store' ORDER BY sort_order")
        store_ids = [r["id"] for r in store_rows]
        asia_mall = store_ids[0] if store_ids else None  # kept for backwards-compat naming

        # Categories + subcategories
        cat_ids: dict[str, uuid.UUID] = {}
        subcat_ids: dict[str, uuid.UUID] = {}

        for p in products:
            if p["category"] not in cat_ids:
                cat_ids[p["category"]] = uuid.uuid4()
            key = f"{p['category']}::{p['subcategory']}"
            if key not in subcat_ids:
                subcat_ids[key] = uuid.uuid4()

        cat_order = {"Мужчинам": 1, "Женщинам": 2, "Аксессуары": 3}
        for name, cid in cat_ids.items():
            await conn.execute(
                "INSERT INTO categories (id, name, slug, sort_order) VALUES ($1, $2, $3, $4)",
                cid, name, slugify(name), cat_order.get(name, 10),
            )
        print(f"Inserted {len(cat_ids)} categories")

        for i, (key, sid) in enumerate(subcat_ids.items()):
            cat_name, subcat_name = key.split("::", 1)
            await conn.execute(
                "INSERT INTO subcategories (id, category_id, name, slug, sort_order) VALUES ($1, $2, $3, $4, $5)",
                sid, cat_ids[cat_name], subcat_name, slugify(f"{cat_name}-{subcat_name}"), i + 1,
            )
        print(f"Inserted {len(subcat_ids)} subcategories")

        # Products
        seen_slugs: set[str] = set()
        seen_skus: set[str] = set()
        inserted = 0
        for p in products:
            if p["price"] <= 0:
                continue

            slug = p["slug"]
            base = slug
            counter = 1
            while slug in seen_slugs:
                counter += 1
                slug = f"{base}-{counter}"
            seen_slugs.add(slug)

            sku = p["sku"]
            if sku:
                if sku in seen_skus:
                    sku = None
                else:
                    seen_skus.add(sku)

            cat_id = cat_ids[p["category"]]
            subcat_id = subcat_ids[f"{p['category']}::{p['subcategory']}"]

            prod_id = uuid.uuid4()
            try:
                await conn.execute(
                    """INSERT INTO products
                       (id, sku, name, slug, description, price, original_price,
                        category_id, subcategory_id, image_url, images, sizes, colors,
                        stock, is_active, is_featured, sort_order)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11::jsonb,$12::jsonb,$13::jsonb,$14,true,$15,0)""",
                    prod_id, sku, p["name"], slug, p["description"],
                    p["price"], p["original_price"],
                    cat_id, subcat_id,
                    p["image_url"], json.dumps(p["images"]),
                    json.dumps(p["sizes"]), json.dumps(p["colors"]),
                    p["stock"], p["original_price"] is not None,
                )
                inserted += 1

                # Seed inventory per-size across all active stores
                if store_ids and p["sizes"] and p["stock"] > 0:
                    per_size = max(1, p["stock"] // len(p["sizes"]))
                    for loc_id in store_ids:
                        for size in p["sizes"]:
                            await conn.execute(
                                """INSERT INTO store_inventory (id, location_id, product_id, size, quantity)
                                   VALUES ($1, $2, $3, $4, $5)
                                   ON CONFLICT (location_id, product_id, size) DO UPDATE SET quantity = EXCLUDED.quantity""",
                                uuid.uuid4(), loc_id, prod_id, size, per_size,
                            )
            except Exception as e:
                print(f"  SKIP {p['name']}: {e}")

        # Stats
        total = await conn.fetchval("SELECT COUNT(*) FROM products")
        inv_rows = await conn.fetchval("SELECT COUNT(*) FROM store_inventory")
        sale = await conn.fetchval("SELECT COUNT(*) FROM products WHERE original_price IS NOT NULL")

        print(f"\n{'=' * 60}")
        print(f"IMPORT COMPLETE")
        print(f"  Products:      {total}")
        print(f"  On sale:       {sale}")
        print(f"  Inventory rows (Asia Mall): {inv_rows}")
        print(f"{'=' * 60}")

    finally:
        await conn.close()


async def main():
    print("Parsing Excel...")
    products = parse_excel()
    print(f"Parsed {len(products)} unique products from {sum(1 for _ in products)} variant rows")

    # Show sample
    print("\nSample products:")
    for p in products[:3]:
        print(f"  {p['name']} — {p['price']} KGS — {len(p['sizes'])} sizes, {len(p['colors'])} colors, {len(p['images'])} imgs, stock={p['stock']}")

    # Save backup
    with open("imported_products.json", "w", encoding="utf-8") as f:
        json.dump(
            [{**p, "price": str(p["price"]), "original_price": str(p["original_price"]) if p.get("original_price") else None} for p in products],
            f, ensure_ascii=False, indent=2,
        )
    print(f"\nBackup saved to imported_products.json")

    await seed(products)


if __name__ == "__main__":
    asyncio.run(main())
